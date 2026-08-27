"""
Unit tests for Excel Migration Documentation generator and API endpoint.
"""

import io
import openpyxl
import pytest
from app.models.job import Job
from app.models.objects import MigrationObject
from app.services.excel_exporter import generate_migration_excel_bytes


def test_excel_exporter_generates_all_5_sheets(db_session):
    # 1. Create a test job
    job = Job(
        id="test-job-excel-1234",
        name="Sales Performance Dossier",
        status="COMPLETE",
        mstr_base_url="https://mstr.enterprise.com/MicroStrategyLibrary",
        mstr_project_id="B7CA92F04B9FAE8D941C3E9B7E0CD754",
        mstr_project_name="Sales Analytics",
        mstr_version="2024.0402",
        tableau_server_url="https://tableau.enterprise.com",
        tableau_site_id="default",
        tableau_target_project="Migrated Executive Dashboards",
        template_version="2024.2",
        security_confidence=1.0,
        financial_kpi_confidence=0.99,
        structural_confidence=1.0,
        visual_confidence=0.98,
        security_parity=True,
    )
    db_session.add(job)

    # 2. Add test objects
    obj_dossier = MigrationObject(
        job_id=job.id,
        mstr_id="dossier-001",
        mstr_type=55,
        type_name="dossier",
        name="Executive Sales Overview",
        status="published",
    )
    obj_cube = MigrationObject(
        job_id=job.id,
        mstr_id="cube-001",
        mstr_type=3,
        type_name="cube",
        name="Sales_Transaction_Cube",
        status="published",
    )
    obj_attr = MigrationObject(
        job_id=job.id,
        mstr_id="attr-001",
        mstr_type=12,
        type_name="attribute",
        name="Region",
        status="published",
    )
    obj_metric = MigrationObject(
        job_id=job.id,
        mstr_id="metric-001",
        mstr_type=4,
        type_name="metric",
        name="Total Direct Sales",
        expression_text="Sum([Direct Sales Amount])",
        tableau_calc="SUM([Direct Sales Amount])",
        tableau_field_name="[Total Direct Sales]",
        translation_method="AST Transpiler",
        confidence=0.99,
        status="published",
    )
    db_session.add_all([obj_dossier, obj_cube, obj_attr, obj_metric])
    db_session.commit()

    # 3. Generate Excel bytes
    excel_bytes = generate_migration_excel_bytes(job.id, db_session)
    assert excel_bytes is not None
    assert len(excel_bytes) > 0

    # 4. Load workbook from bytes and verify sheets
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
    sheet_names = wb.sheetnames
    assert len(sheet_names) == 5
    assert sheet_names == [
        "Overview & KPIs",
        "MSTR Source Metadata",
        "Metric & Logic Translation",
        "Visual & Layout Mapping",
        "Execution & Audit Trail",
    ]

    # Verify Sheet 1 (Overview)
    ws1 = wb["Overview & KPIs"]
    assert "MicroStrategy" in str(ws1.cell(row=1, column=1).value)
    assert ws1.cell(row=4, column=2).value == "Sales Performance Dossier"

    # Verify Sheet 2 (Metadata)
    ws2 = wb["MSTR Source Metadata"]
    assert ws2.cell(row=1, column=2).value == "MSTR Object ID (GUID)"
    assert ws2.max_row >= 5  # Header + 4 objects

    # Verify Sheet 3 (Calculations)
    ws3 = wb["Metric & Logic Translation"]
    assert ws3.cell(row=1, column=2).value == "MSTR Metric Name"
    assert ws3.cell(row=1, column=4).value == "Tableau Calculated Field Name"

    # Verify no sheet contains any "Confidence" column header or scorecard
    for s_name in sheet_names:
        ws = wb[s_name]
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val:
                    assert "confidence" not in str(val).lower(), f"Unexpected 'confidence' found in sheet {s_name}: {val}"


def test_excel_export_api_endpoint(client, db_session):
    job = Job(
        id="api-test-job-excel",
        name="Claims Migration",
        status="COMPLETE",
        mstr_base_url="https://mstr.test.com/MicroStrategyLibrary",
        mstr_project_id="PROJ123",
    )
    db_session.add(job)
    db_session.commit()

    response = client.get(f"/api/v1/jobs/{job.id}/export/excel")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "Content-Disposition" in response.headers
    assert ".xlsx" in response.headers["Content-Disposition"]
    assert len(response.content) > 1000
